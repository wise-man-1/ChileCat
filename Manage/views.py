"""管理视图"""
import logging
import datetime
import time
from django.http import JsonResponse, HttpResponse
from openpyxl import load_workbook
from rest_framework.views import APIView
from Apps.Permission.utils import expand_permission
from Apps.User.models import UserInfo, Grade, College, User, StudentInfo, WholeGrade, TeacherForGrade
from Apps.Life.models import Building, Floor, Room, StuInRoom
from Apps.Activity.models import TaskRecord
from django.template import loader
from Apps.Permission.models import *
# from django.contrib.contenttypes.models import ContentType
# from docxtpl import DocxTemplate
from .tests import *
from django.core import serializers
from Apps.Activities.Attendance.Entity.ser import TaskRecordSerializer

# from rest_framework.authtoken.models import Token

logger = logging.getLogger(__name__)


class Test(APIView):
    """后台接口调用"""

    def get(self, request):
        """测试接口"""
        print("GET参数:", self.request.query_params)
        print("request data:", self.request.data)
        print("TOKEN:", self.request.META.get("HTTP_TOKEN"))
        print('测试接口')

        # 当销假人为空的时候为未销假

        # grades = Grade.objects.all()
        # for grade in grades:
        #     relate = TeacherForGrade.objects.get_or_create(user=User.objects.get(username="admin"), grade=grade)
        #     print("admin->", relate)
        # user = User.objects.get(username="206500000")
        # WholeGrade.objects.create(name="20级",user=user)
        # data = serializers.serialize("json", Grade.objects.all())
        # print(data)

        # grades = Grade.objects.all()
        # for grade in grades:
        #     if grade.name[0:3] == "195":
        #         grade.whole_grade = WholeGrade.objects.get(id=1)
        #         grade.save()
        #     elif grade.name[0:3] == "206":
        #         grade.whole_grade = WholeGrade.objects.get(id=2)
        #         grade.save()

        # 用户名小写,密码默认123456
        # users = User.objects.all()
        # for user in users:
        #     if user.username != user.username.lower():
        #         user.username = user.username.lower()
        #     # user.set_password("123456")
        #     # user.save()
        #     if len(user.username) == 8 and user.username[0:3] == "206":  # 205
        #         print(user.username)

        #
        # 班级改小写
        # grades = Grade.objects.all()
        # for grade in grades:
        #     if grade.name != grade.name.lower():
        #         grade.name = grade.name.lower()
        #         grade.save()
        #
        # print("创建班主任:")
        # grades = Grade.objects.all()
        # for grade in grades:
        #     grade_name = grade.name
        #     teacher_username = grade_name + "00"
        #     teacher, flag = User.objects.get_or_create(username=teacher_username)
        #     teacher.set_password("123456")
        #     teacher.save()
        #     UserInfo.objects.get_or_create(user=teacher, name=grade_name+"班老师", identity="teacher")
        #     print(TeacherForGrade.objects.get_or_create(user=teacher, grade=grade))

        # import_stu_data("leaksfile//副本智慧交通学院学生寝室信息表（全).xlsx")
        # add_student("leaksfile//学生导入表.xlsx")
        return JsonResponse({"message": "task"})


def add_student(file):
    """针对寝室导入表"""
    work_book = load_workbook(file)
    log = open(file + ".log", "w")
    for sheet in work_book:
        room_name = ""
        for info in sheet.values:
            print(info)
            if info[0] == "寝室".strip():
                continue
            if info[0]:
                room_name = str(info[0])[0:info[0].find("#") + 3 + 1]
            name = ''.join(re.findall('[\u4e00-\u9fa5]', str(info[1]))).strip()
            grade = str(info[2]).strip()
            stu_id = str(info[3]).strip()
            tel = str(info[4]).strip()
            print(room_name, name, grade, stu_id, tel)
            try:
                user = User.objects.get_or_create(username=stu_id)
                UserInfo.objects.get_or_create(user=user[0], name=name, tel=tel)
                college = College.objects.get_or_create(name="智慧交通学院")
                grade = Grade.objects.get_or_create(name=grade, college=college[0])
                StudentInfo.objects.get_or_create(user=user[0], grade=grade[0])
                put_stu_room(user[0], room_name, log)
            except User.DoesNotExist:
                print(name, grade, stu_id, tel)
            except UserInfo.DoesNotExist:
                print(name, grade, stu_id, tel)


def put_stu_room(stu, room, log):
    """把学生放入寝室,注意第二个参数目前只支持xx#xxx形式"""
    student = stu
    history = StuInRoom.objects.filter(room=search_room(room), student=student)
    if history.exists():
        print("学生", student.userinfo.name, "->", history.first(), "已存在")
    else:
        room = search_room(room)
        stu_in_room = StuInRoom.objects.get_or_create(room=room, student=student)
        print("记录(", "学号:", student.username, "姓名:", student.userinfo.name, "->", stu_in_room[0], "寝室)已创建")


def search_room(room_info):
    """xx#xxx解析"""
    building_name = room_info.strip().split("#")[0].strip()
    floor = room_info.strip().split("#")[1].strip()[0].strip()
    room = room_info.strip().split("#")[1].strip()[1:3].strip()
    building, flag = Building.objects.get_or_create(name=building_name)
    floor, flag = Floor.objects.get_or_create(building=building, name=floor)
    room, flag = Room.objects.get_or_create(name=room, floor=floor)
    return room


def create_class(class_name, college_name):
    """创建班级"""
    if not Grade.objects.filter(name=class_name).exists():
        if not College.objects.filter(name=college_name).exists():
            college = College(name=college_name)
            college.save()
            print("分院:", college, "创建.")
        college = College.objects.get(name=college_name)
        grade = Grade(name=class_name, college=college)
        grade.save()
        print("班级", grade, "创建")
        return grade
    print("班级已存在,无需创建!")
    return None


def import_stu_data(file):
    """针对寝室表"""
    work_book = load_workbook(file)
    log = open(file + ".log", "w")
    for sheet in work_book:
        room_name = ""
        for info in sheet.values:
            if info[0] == "寝室日".strip():
                continue
            if info[0]:
                room_name = str(info[0])[0:info[0].find("#") + 3 + 1]
            name = ''.join(re.findall('[\u4e00-\u9fa5]', str(info[2]))).strip()
            grade = str(info[6]).strip()
            stu_id = str(info[7]).strip()
            tel = str(info[8]).strip()
            try:
                user = User.objects.get_or_create(username=stu_id)
                UserInfo.objects.get_or_create(user=user[0], name=name, tel=tel)
                college = College.objects.get_or_create(name="智慧交通学院")
                grade = Grade.objects.get_or_create(name=grade, college=college[0])
                StudentInfo.objects.get_or_create(user=user[0], grade=grade[0])
                put_stu_room(user[0], room_name, log)
            except User.DoesNotExist:
                print(name, grade, stu_id, tel)
            except UserInfo.DoesNotExist:
                print(name, grade, stu_id, tel)


class ApiPer(APIView):
    """API权限生成"""

    def get(self, request):
        """test"""
        ret = {'message': 'message', 'code': 2000}

        # expand_permission.group_clean('life_admin')
        # expand_permission.user_admin_clean(['19510144','19510143'])
        # expand_permission.init_api_permissions()
        return JsonResponse(ret)


class Index(APIView):
    """"""

    def get(self, request):
        api_permission_list = ApiPermission.objects.filter(
            permission__codename="/api/life/idcode:GET").values_list('permission__codename', flat=True)
        urls = []
        id_code = "api/life/idcode"
        for api in api_permission_list:
            ap = str(api).split(":")
            urls.append(ap[0])
        template = loader.get_template('Manage/index.html')
        context = {
            'display_list': api_permission_list,
            'id_code': id_code,
        }
        return HttpResponse(template.render(context, request))


# 用户组初始化
def group_init(request):
    # group2 = ['ask_admin','life_admin']

    # 待添加进用户组的权限
    permissions = [
        '/api/life/switchknowing:PUT',
        '/api/life/switchknowing:POST',
        '/api/life/idcode:PUT',
        '/api/life/recordsearch:PUT',
        '/api/life/studentleak:PUT',
    ]

    # 用户组
    name = 'life_admin'

    # 用户
    users = [
        '19510140',
    ]

    expand_permission.group_init(name)
    # expand_permission.group_add_permission(name,permissions)
    expand_permission.group_add_user(name, users)

    return JsonResponse(2000, safe=False)


# 寝室调换

def dormitory_exchange(request):
    ret = {'message': 'message', 'code': 2000}
    # 待调换的数据
    data = {
        '20653w213': ['3', '4', '22'],
        '19530139': ['3', '3', '04'],
        '19530116': ['3', '3', '03'],
        '1853w115': ['3', '2', '14']
    }
    for item in data:
        try:
            user = User.objects.get(username=item)
            # 新寝室
            b = Building.objects.get(name=data[item][0])
            f = Floor.objects.get(name=data[item][1], building=b)
            r = Room.objects.get(name=data[item][2], floor=f)
            print('学生：', user, '旧寝室：', user.stu_in_room.all()[0], '待换寝室：', r)
            StuInRoom.objects.filter(student=user).update(room=r)
        except:
            print('调换失败', item)

    return JsonResponse(ret)


def add_user():
    L = [
        ['195401', '19540140'],
        ['195303', '19530338'],
        ['195303', '19530345'],
    ]
